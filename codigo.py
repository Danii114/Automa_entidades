import pandas as pd
import requests
import smtplib
import re
import numpy as np
from datetime import date, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =========================================================
# CONFIGURACIÓN
# =========================================================

EMAIL_SENDER = "danimorav05@gmail.com"
EMAIL_PASSWORD = "exum hoyz gcvy nsli"

EMAIL_TO = [
    "danimorav05@gmail.com",
    "karenselennemoreno@gmail.com"
]
URL_SECOP1 = "https://www.datos.gov.co/resource/f789-7hwg.json"
URL_SECOP2 = "https://www.datos.gov.co/resource/jbjy-vk9h.json"

DIAS_ATRAS = 8

fecha_fin = date.today()
fecha_inicio = fecha_fin - timedelta(days=DIAS_ATRAS)
fecha_str = fecha_fin.strftime("%Y-%m-%d")

ARCHIVO_EXCEL = f"reporte_secop_i_ii_{fecha_str}.xlsx"

# NIT correctos para SECOP I
NITS_SECOP1 = [
    "899999001",  # Ministerio de Educación
    "899999083",  # DAPRE / Fondo Paz
    "830114475",  # Ministerio del Interior
    "599999014",  # Ministerio de Salud
    "900002583",  # RTVC
    "800131648"   # FUTIC en SECOP I
]

# NIT para SECOP II
NITS_SECOP2 = [
    "899999001",
    "899999083",
    "830114475",
    "900474727",
    "900002583",
    "8001316486"  # FUTIC en SECOP II
]

NITS_SQL_SECOP1 = ", ".join([f"'{nit}'" for nit in NITS_SECOP1])
NITS_SQL_SECOP2 = ", ".join([f"'{nit}'" for nit in NITS_SECOP2])

# =========================================================
# FUNCIONES
# =========================================================

def safe_float(x):
    try:
        return float(x)
    except:
        return 0.0

def extraer_url(x):
    if pd.isna(x):
        return ""
    x = str(x)
    m = re.search(r"https?://[^\s]+", x)
    return m.group(0) if m else (x if x.startswith("http") else "")

def limpiar_valor(val):
    if val is None:
        return ""
    if isinstance(val, float) and np.isnan(val):
        return ""
    return str(val)

def descargar_socrata(url, query_base, limit=50000):
    datos = []
    offset = 0

    while True:
        query = f"{query_base}\nLIMIT {limit} OFFSET {offset}"
        r = requests.get(url, params={"$query": query})

        print(f"Status: {r.status_code} | offset: {offset}")

        if r.status_code != 200:
            print(r.text[:1000])
            break

        bloque = r.json()

        if not bloque:
            break

        datos.extend(bloque)

        if len(bloque) < limit:
            break

        offset += limit

    return pd.DataFrame(datos)

# =========================================================
# DESCARGAR SECOP I
# Fecha: fecha_de_firma_del_contrato
# Estado: CONVOCADO
# =========================================================

print("\n📥 Descargando SECOP I...")

query_secop1 = f"""
SELECT
    nombre_entidad,
    nit_de_la_entidad,
    detalle_del_objeto_a_contratar,
    modalidad_de_contratacion,
    cuantia_proceso,
    plazo_de_ejec_del_contrato,
    fecha_de_firma_del_contrato,
    estado_del_proceso,
    ruta_proceso_en_secop_i
WHERE nit_de_la_entidad IN ({NITS_SQL_SECOP1})
AND upper(estado_del_proceso) = 'CONVOCADO'
AND fecha_de_firma_del_contrato >= '{fecha_inicio}T00:00:00'
AND fecha_de_firma_del_contrato <= '{fecha_fin}T23:59:59'
"""

df_secop1 = descargar_socrata(URL_SECOP1, query_secop1)

print(f"SECOP I registros: {len(df_secop1)}")

# =========================================================
# DESCARGAR SECOP II
# Fecha: fecha_de_firma
# Sin filtro por estado
# =========================================================

print("\n📥 Descargando SECOP II...")

query_secop2 = f"""
SELECT
    nombre_entidad,
    nit_entidad,
    id_contrato,
    descripcion_del_proceso,
    valor_del_contrato,
    fecha_de_firma,
    estado_contrato,
    duraci_n_del_contrato,
    modalidad_de_contratacion,
    proveedor_adjudicado,
    urlproceso
WHERE nit_entidad IN ({NITS_SQL_SECOP2})
AND fecha_de_firma >= '{fecha_inicio}T00:00:00'
AND fecha_de_firma <= '{fecha_fin}T23:59:59'
"""

df_secop2 = descargar_socrata(URL_SECOP2, query_secop2)

print(f"SECOP II registros: {len(df_secop2)}")

# =========================================================
# LIMPIEZA
# =========================================================

COLS_S1 = [
    "nombre_entidad",
    "nit_de_la_entidad",
    "detalle_del_objeto_a_contratar",
    "modalidad_de_contratacion",
    "cuantia_proceso",
    "plazo_de_ejec_del_contrato",
    "fecha_de_firma_del_contrato",
    "estado_del_proceso",
    "ruta_proceso_en_secop_i"
]

COLS_S2 = [
    "nombre_entidad",
    "nit_entidad",
    "id_contrato",
    "descripcion_del_proceso",
    "valor_del_contrato",
    "fecha_de_firma",
    "estado_contrato",
    "duraci_n_del_contrato",
    "modalidad_de_contratacion",
    "proveedor_adjudicado",
    "valor_diario_contrato_cop",
    "urlproceso"
]

if df_secop1.empty:
    df_secop1 = pd.DataFrame(columns=COLS_S1)

if df_secop2.empty:
    df_secop2 = pd.DataFrame(columns=COLS_S2)

df_secop1["cuantia_proceso"] = pd.to_numeric(
    df_secop1.get("cuantia_proceso", pd.Series(dtype=str)).astype(str)
    .str.replace(",", "", regex=False)
    .str.strip(),
    errors="coerce"
).fillna(0)

df_secop2["valor_del_contrato"] = pd.to_numeric(
    df_secop2.get("valor_del_contrato", pd.Series(dtype=str)).astype(str)
    .str.replace(",", "", regex=False)
    .str.strip(),
    errors="coerce"
).fillna(0)

df_secop2["cuantia_proceso"] = df_secop2["valor_del_contrato"]
df_secop2["detalle_del_objeto_a_contratar"] = df_secop2.get("descripcion_del_proceso", "")

#agregando columnas 
# =========================================================
# VALOR DIARIO DEL CONTRATO - SECOP II
# La columna duraci_n_del_contrato ya viene en días
# =========================================================

# Extraer el número de días desde textos como "207 Día(s)"
df_secop2["duracion_contrato_dias"] = (
    df_secop2["duraci_n_del_contrato"]
    .astype(str)
    .str.extract(r"(\d+)")
    .astype(float)
)

# Evitar división por cero
df_secop2.loc[
    df_secop2["duracion_contrato_dias"] <= 0,
    "duracion_contrato_dias"
] = np.nan

# Calcular valor diario en pesos colombianos
df_secop2["valor_diario_contrato_cop"] = (
    df_secop2["valor_del_contrato"] /
    df_secop2["duracion_contrato_dias"]
).round(2)

# Verificación rápida
print("\nVerificación SECOP II valor diario:")
print(
    df_secop2[
        [
            "duraci_n_del_contrato",
            "duracion_contrato_dias",
            "valor_del_contrato",
            "valor_diario_contrato_cop"
        ]
    ].head(10)
)
# =========================================================
# EXCEL FORMATEADO
# =========================================================

print("\n📊 Generando Excel...")

borde = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

ANCHOS = {
    "nombre_entidad": 35,
    "nit_de_la_entidad": 18,
    "nit_entidad": 18,
    "detalle_del_objeto_a_contratar": 55,
    "descripcion_del_proceso": 65,
    "modalidad_de_contratacion": 25,
    "cuantia_proceso": 18,
    "valor_del_contrato": 18,
    "plazo_de_ejec_del_contrato": 18,
    "fecha_de_firma_del_contrato": 24,
    "fecha_de_firma": 24,
    "estado_del_proceso": 18,
    "estado_contrato": 18,
    "duraci_n_del_contrato": 18,
    "proveedor_adjudicado": 18,
    "id_contrato": 20,
    "valor_diario_contrato_cop": 20,
    "ruta_proceso_en_secop_i": 18,
    "urlproceso": 18,
}

def escribir_hoja(ws, df, cols, url_col, header_hex, par_hex):
    cols_ok = [c for c in cols if c in df.columns]
    df_ex = df[cols_ok].copy()

    for money_col in ["cuantia_proceso", "valor_del_contrato"]:
        if money_col in df_ex.columns:
            df_ex[money_col] = df_ex[money_col].apply(
                lambda x: f"${float(x):,.0f}" if pd.notnull(x) and str(x) != "" else ""
            )

    urls = df_ex[url_col].apply(extraer_url).tolist() if url_col in df_ex.columns else [""] * len(df_ex)
    url_idx = cols_ok.index(url_col) + 1 if url_col in cols_ok else None

    df_ex = df_ex.apply(lambda col: col.map(limpiar_valor))

    ws.append(cols_ok)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=header_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borde

    ws.row_dimensions[1].height = 28

    for i, (row_data, url_val) in enumerate(zip(df_ex.itertuples(index=False), urls), start=2):
        ws.append(list(row_data))

        color = par_hex if i % 2 == 0 else "FFFFFF"

        for cell in ws[i]:
            cell.fill = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = borde

        if url_idx and url_val and url_val.startswith("http"):
            c = ws.cell(row=i, column=url_idx)
            c.hyperlink = url_val
            c.value = "Ver proceso"
            c.font = Font(color="0563C1", underline="single", size=10)

        ws.row_dimensions[i].height = 36

    for i, col_name in enumerate(cols_ok, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ANCHOS.get(col_name, 20)

    ws.freeze_panes = "A2"

wb = Workbook()

ws1 = wb.active
ws1.title = "SECOP I"
escribir_hoja(ws1, df_secop1, COLS_S1, "ruta_proceso_en_secop_i", "1F4E79", "EEF4FF")

ws2 = wb.create_sheet("SECOP II")
escribir_hoja(ws2, df_secop2, COLS_S2, "urlproceso", "166534", "ECFDF5")

excel_ok = False

try:
    wb.save(ARCHIVO_EXCEL)
    excel_ok = True
    print(f"Excel guardado: {ARCHIVO_EXCEL}")
except Exception as e:
    print(f"Error guardando Excel: {e}")

# =========================================================
# HTML CORREO
# =========================================================

def metrics_block(df):
    total = len(df)
    entidades = df["nombre_entidad"].nunique() if "nombre_entidad" in df.columns else 0
    v_total = float(df["cuantia_proceso"].sum()) if "cuantia_proceso" in df.columns else 0
    v_prom = float(df["cuantia_proceso"].mean()) if total > 0 else 0
    return total, entidades, v_total, v_prom

def build_top5_rows(df, url_col):
    if df.empty:
        return "<tr><td colspan='4' style='padding:10px;text-align:center;color:#999;font-size:12px;'>Sin registros en el período</td></tr>"

    top5 = df.sort_values("cuantia_proceso", ascending=False).head(5)

    rows = ""

    for _, row in top5.iterrows():
        link = extraer_url(str(row.get(url_col, "")))
        link_html = (
            f"<a href='{link}' style='color:#1F4E79;font-weight:bold;text-decoration:none;'>&#128279; Ver</a>"
            if link else "&#8212;"
        )

        cuantia = safe_float(row.get("cuantia_proceso", 0))
        desc = str(row.get("detalle_del_objeto_a_contratar") or "")[:65]
        entidad = str(row.get("nombre_entidad", ""))[:38]

        rows += f"""
        <tr>
          <td style="padding:7px 6px;border-bottom:1px solid #f0f0f0;font-size:12px;color:#222;">{entidad}</td>
          <td style="padding:7px 6px;border-bottom:1px solid #f0f0f0;font-size:12px;color:#555;">{desc}&#8230;</td>
          <td style="padding:7px 6px;border-bottom:1px solid #f0f0f0;font-size:12px;color:#1a7f4b;font-weight:bold;white-space:nowrap;">${cuantia:,.0f}</td>
          <td style="padding:7px 6px;border-bottom:1px solid #f0f0f0;font-size:12px;text-align:center;">{link_html}</td>
        </tr>
        """

    return rows

def seccion_html(label, text_color, bg_color, border_color,
                 total, entidades, v_total, v_prom, top5_rows):

    vp = f"${v_prom:,.0f}" if v_prom else "&#8212;"

    return f"""
  <div style="background:{bg_color};padding:6px 20px;font-size:11px;font-weight:700;
              color:{text_color};border-top:1px solid {border_color};border-bottom:1px solid {border_color};">
    &#9679;&nbsp; {label}
  </div>

  <table width="100%" cellpadding="0" cellspacing="0" style="border-bottom:1px solid #e5e7eb;">
    <tr>
      <td width="25%" style="padding:12px 6px;text-align:center;border-right:1px solid #e5e7eb;">
        <p style="margin:0;font-size:24px;font-weight:700;color:{text_color};">{total}</p>
        <p style="margin:2px 0 0;font-size:10px;color:#888;">REGISTROS</p>
      </td>

      <td width="25%" style="padding:12px 6px;text-align:center;border-right:1px solid #e5e7eb;">
        <p style="margin:0;font-size:24px;font-weight:700;color:{text_color};">{entidades}</p>
        <p style="margin:2px 0 0;font-size:10px;color:#888;">ENTIDADES</p>
      </td>

      <td width="25%" style="padding:12px 6px;text-align:center;border-right:1px solid #e5e7eb;">
        <p style="margin:0;font-size:14px;font-weight:700;color:#1a7f4b;">${v_total:,.0f}</p>
        <p style="margin:2px 0 0;font-size:10px;color:#888;">VALOR TOTAL</p>
      </td>

      <td width="25%" style="padding:12px 6px;text-align:center;">
        <p style="margin:0;font-size:14px;font-weight:700;color:#1a7f4b;">{vp}</p>
        <p style="margin:2px 0 0;font-size:10px;color:#888;">VALOR PROMEDIO</p>
      </td>
    </tr>
  </table>

  <div style="padding:14px 20px 16px;">
    <p style="margin:0 0 8px;font-size:13px;font-weight:700;color:#1F4E79;">
      &#128176;&nbsp; Top 5 por valor
    </p>

    <table width="100%" cellpadding="0" cellspacing="0">
      <thead>
        <tr style="background:#f8fafc;">
          <th style="padding:6px;text-align:left;font-size:10px;color:#888;">ENTIDAD</th>
          <th style="padding:6px;text-align:left;font-size:10px;color:#888;">DESCRIPCIÓN</th>
          <th style="padding:6px;text-align:left;font-size:10px;color:#888;">VALOR</th>
          <th style="padding:6px;text-align:center;font-size:10px;color:#888;">LINK</th>
        </tr>
      </thead>
      <tbody>{top5_rows}</tbody>
    </table>
  </div>
  """

t1, e1, vt1, vp1 = metrics_block(df_secop1)
t2, e2, vt2, vp2 = metrics_block(df_secop2)

rows1 = build_top5_rows(df_secop1, "ruta_proceso_en_secop_i")
rows2 = build_top5_rows(df_secop2, "urlproceso")

total_global = t1 + t2
valor_global = vt1 + vt2

s1 = seccion_html(
    "SECOP I &#8212; Estado CONVOCADO",
    "#1e40af",
    "#EFF6FF",
    "#bfdbfe",
    t1, e1, vt1, vp1, rows1
)

s2 = seccion_html(
    "SECOP II &#8212; Fecha de firma",
    "#166534",
    "#F0FDF4",
    "#bbf7d0",
    t2, e2, vt2, vp2, rows2
)

resumen_html = f"""
<html>
<body style="font-family:Arial,sans-serif;background:#f1f5f9;padding:16px;margin:0;">
<table width="100%" cellpadding="0" cellspacing="0">
<tr>
<td align="center">

<table width="680" cellpadding="0" cellspacing="0"
       style="background:#fff;border-radius:8px;overflow:hidden;border:1px solid #e2e8f0;">

<tr>
<td style="background:#1F4E79;padding:20px 24px;">
  <p style="margin:0;font-size:19px;font-weight:700;color:#fff;">
    &#128204;&nbsp; SECOP &#8212; I-II
  </p>
  <p style="margin:4px 0 0;font-size:12px;color:#93c5fd;">
    Periodo: {fecha_inicio.strftime('%Y-%m-%d')} &rarr; {fecha_str}
    &nbsp;&middot;&nbsp; {date.today().strftime("%d/%m/%Y")}
    &nbsp;&nbsp;|&nbsp;&nbsp;
    <strong style="color:#fff;">{total_global} registros</strong>
    &nbsp;&middot;&nbsp;
    <strong style="color:#fff;">${valor_global:,.0f}</strong>
  </p>
</td>
</tr>

<tr>
<td>
  <table width="100%" cellpadding="0" cellspacing="0"
         style="border-bottom:3px solid #1F4E79;">
    <tr>
      <td width="50%" style="padding:9px;text-align:center;background:#1F4E79;
          font-size:12px;font-weight:700;color:#fff;">
        SECOP I &nbsp;&middot;&nbsp; {t1} registros
      </td>
      <td width="50%" style="padding:9px;text-align:center;background:#f8fafc;
          font-size:12px;font-weight:700;color:#1F4E79;">
        SECOP II &nbsp;&middot;&nbsp; {t2} registros
      </td>
    </tr>
  </table>
</td>
</tr>

<tr><td>{s1}</td></tr>
<tr><td style="height:6px;background:#f1f5f9;"></td></tr>
<tr><td>{s2}</td></tr>

<tr>
<td style="background:#f8fafc;padding:10px 20px;text-align:center;
           border-top:1px solid #e2e8f0;">
  <p style="margin:0;font-size:11px;color:#94a3b8;">
    &#128206;&nbsp; Excel adjunto con las dos hojas completas
  </p>
</td>
</tr>

</table>

</td>
</tr>
</table>
</body>
</html>
"""

# =========================================================
# ENVÍO CORREO
# =========================================================

print("\n📧 Enviando correo...")

msg = MIMEMultipart("alternative")
msg["Subject"] = f"SECOP I-II entidades · {fecha_str} · {total_global} registros (I:{t1} / II:{t2})"
msg["From"] = EMAIL_SENDER
msg["To"] = ", ".join(EMAIL_TO)

msg.attach(MIMEText(resumen_html, "html"))

if excel_ok:
    try:
        with open(ARCHIVO_EXCEL, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())

        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f"attachment; filename={ARCHIVO_EXCEL}"
        )

        msg.attach(part)

    except Exception as e:
        print(f"No se pudo adjuntar Excel: {e}")

try:
    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls()
        server.login(EMAIL_SENDER, EMAIL_PASSWORD)
        server.sendmail(EMAIL_SENDER, EMAIL_TO, msg.as_string())

    print("Correo enviado correctamente")

except Exception as e:
    print(f"Error al enviar correo: {e}")

print("\nPipeline finalizado")
